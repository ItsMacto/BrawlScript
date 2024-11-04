import requests
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv

load_dotenv()

API_KEY = os.getenv("API_KEY")
CLUB_TAG = os.getenv("CLUB_TAG")


CLUB_URL = f"https://api.brawlstars.com/v1/clubs/{CLUB_TAG.replace('#', '%23')}"
PLAYER_STATS_URL = "https://api.brawlstars.com/v1/players/{playerTag}"

headers = {
    "Authorization": f"Bearer {API_KEY}",
    "Accept": "application/json"
}

def fetch_club_info():
    """Fetches general information about the club."""
    response = requests.get(CLUB_URL, headers=headers)
    response.raise_for_status()
    return response.json()

def fetch_club_members():
    """Fetches the club members' information."""
    club_data = fetch_club_info()
    members = {member['name']: member['trophies'] for member in club_data.get('members', [])}
    return members, club_data

def update_club_sheet(workbook, club_data):
    """Updates or creates a sheet for club-level statistics."""
    if "Club Stats" not in workbook.sheetnames:
        sheet = workbook.create_sheet(title="Club Stats")
        sheet["A1"] = "Date"
        sheet["B1"] = "Total Trophies"
        sheet["C1"] = "Average Trophies"
        sheet["D1"] = "Total Members"
    else:
        sheet = workbook["Club Stats"]

    current_date = datetime.now().strftime("%Y-%m-%d")
    new_row = sheet.max_row + 1
    total_trophies = club_data.get("trophies", 0)
    total_members = len(club_data.get("members", []))
    avg_trophies = total_trophies / total_members if total_members else 0

    sheet[f"A{new_row}"] = current_date
    sheet[f"B{new_row}"] = total_trophies
    sheet[f"C{new_row}"] = avg_trophies
    sheet[f"D{new_row}"] = total_members

def update_member_sheet(workbook, members):
    """Updates the sheet with individual member statistics, adding columns weekly."""
    if "Member Stats" not in workbook.sheetnames:
        sheet = workbook.create_sheet(title="Member Stats")
        sheet["A1"] = "Member Name"
    else:
        sheet = workbook["Member Stats"]

    last_date_cell = sheet.cell(row=1, column=sheet.max_column)
    last_date = last_date_cell.value
    current_date = datetime.now().strftime("%Y-%m-%d")

    if isinstance(last_date, str):
        try:
            last_date = datetime.strptime(last_date, "%Y-%m-%d")
        except ValueError:
            last_date = None

    if last_date is None or (datetime.strptime(current_date, "%Y-%m-%d") - last_date) >= timedelta(days=7):
        new_column = sheet.max_column + 1
        sheet.cell(row=1, column=new_column, value=current_date)

        for row in range(2, sheet.max_row + 1):
            member_name = sheet.cell(row=row, column=1).value

            if member_name in members:
                trophies_now = members[member_name]
                trophies_last_week = sheet.cell(row=row, column=new_column - 1).value if last_date else None
                if trophies_last_week is not None:
                    trophies_gained = trophies_now - trophies_last_week
                else:
                    trophies_gained = trophies_now  # Initial entry if no prior data

                cell = sheet.cell(row=row, column=new_column, value=trophies_now)

                if trophies_last_week is None:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif trophies_gained >= 300:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 225 <= trophies_gained < 300:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                sheet.cell(row=row, column=new_column, value="Left")

        for member_name, trophies in members.items():
            if not any(sheet.cell(row=row, column=1).value == member_name for row in range(2, sheet.max_row + 1)):
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=member_name)
                sheet.cell(row=row, column=new_column, value=trophies)
                sheet.cell(row=row, column=new_column).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

def main():
    try:
        workbook = openpyxl.load_workbook("club_trophies.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    members, club_data = fetch_club_members()
    update_club_sheet(workbook, club_data)
    update_member_sheet(workbook, members)
    workbook.save("club_trophies.xlsx")
    print("Data saved successfully.")

if __name__ == "__main__":
    main()